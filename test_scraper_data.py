import unittest
from unittest.mock import MagicMock, patch
from typing import Any
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from scrape_data import get_driver, write_file, scrape_website
from selenium.webdriver.common.by import By

class TestScraperFunctions(unittest.TestCase):

    def create_mock_excel_file(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Establishment Name (DECRS)'
        ws['B1'] = 'DUNS (DECRS)'
        ws['C1'] = 'Business Operations (DECRS)'
        ws['D1'] = 'Registration Expiration Date (DECRS)'
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def create_excel_for_write_test(self):
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            ws = wb["Sheet"]
            ws.title = "Full Query"
        else:
            ws = wb.create_sheet("Full Query")

        ws.append(['Col1', 'Col2'])
        ws.append(['', 'Establishment 1'])
        ws.append(['Establishment Name (DECRS)', ''])
        ws.append(['DUNS (DECRS)', ''])
        ws.append(['Business Operations (DECRS)', ''])
        ws.append(['Registration Expiration Date (DECRS)', ''])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, wb

    def test_write_file_writes_correct_data(self):
        excel_file, wb = self.create_excel_for_write_test()
        ws = wb.active

        df = pd.DataFrame([
            ['', 'Establishment 1'],
            ['Establishment Name (DECRS)', ''],
            ['DUNS (DECRS)', ''],
            ['Business Operations (DECRS)', ''],
            ['Registration Expiration Date (DECRS)', '']
        ], columns=['Col1', 'Col2'])

        with patch('scrape_data.convert_excel_to_csv', return_value="mock.csv"), \
             patch('pandas.read_csv', return_value=df):

            write_file("mock.xlsx", wb, 0, "Test Name", "123456789", "Consulting", "2025-12-31", "address")

            values = [ws.cell(row=i, column=2).value for i in range(3, 7)]
            self.assertEqual(values, ["Test Name", "123456789", "Consulting", "2025-12-31"])

    @patch('scrape_data.fuzz.partial_ratio', return_value=90)
    @patch('scrape_data.write_file')
    @patch('scrape_data.webdriver.Chrome')
    def test_write_file_called_when_match_found(self, mock_chrome, mock_write_file, mock_fuzz):
        mock_driver = MagicMock()
        mock_chrome.return_value = mock_driver

        row = MagicMock()
        row.find_element.side_effect = lambda by, name=None: MagicMock(text="Test Address")

        table = MagicMock()
        table.find_elements.return_value = [row]
        mock_driver.find_element.return_value = table

        excel_file = self.create_mock_excel_file()
        links = ['https://dps.fda.gov/decrs/searchresult?type=Merck+Sharp+&+Dohme+LLC']
        establishments = {0: 'Test Address'}
        search_param = 'address'

        scrape_website(excel_file, links, establishments, search_param)
        mock_write_file.assert_called_once()

    @patch('scrape_data.fuzz.partial_ratio', return_value=50)
    @patch('scrape_data.write_file')
    @patch('scrape_data.webdriver.Chrome')
    def test_write_file_not_called_when_no_match(self, mock_chrome, mock_write_file, mock_fuzz):
        mock_driver = MagicMock()
        mock_chrome.return_value = mock_driver

        row = MagicMock()
        row.find_element.side_effect = lambda by, name=None: MagicMock(text="Some Other Address")

        table = MagicMock()
        table.find_elements.return_value = [row]
        mock_driver.find_element.return_value = table

        excel_file = self.create_mock_excel_file()
        links = ['https://dps.fda.gov/decrs/searchresult?type=Merck+Sharp+&+Dohme+LLC']
        establishments = {0: 'Unrelated Address'}
        search_param = 'address'

        scrape_website(excel_file, links, establishments, search_param)
        mock_write_file.assert_not_called()

    @patch('scrape_data.ChromeDriverManager')
    @patch('scrape_data.webdriver.Chrome')
    def test_get_driver_returns_chrome_driver(self, mock_chrome, mock_manager):
        mock_manager.return_value.install.return_value = 'mock_path'
        mock_instance = MagicMock()
        mock_chrome.return_value = mock_instance

        driver = get_driver()
        mock_chrome.assert_called_once()
        self.assertEqual(driver, mock_instance)

    @patch('selenium.webdriver.Chrome')
    def test_scrape_website_returns_excel_buffer(self, mock_chrome):
        mock_driver = MagicMock()
        mock_chrome.return_value = mock_driver
        mock_driver.find_element.return_value.text = "Test Data"

        mock_row = MagicMock()
        mock_row.find_element.return_value.text = "Test Address"

        table = MagicMock()
        table.find_elements.return_value = [mock_row]
        mock_driver.find_element.return_value = table

        excel_file = self.create_mock_excel_file()
        links = ['https://dps.fda.gov/decrs/searchresult?type=Merck+Sharp+&+Dohme+LLC']
        establishments = {1: 'Test Address'}
        search_param = 'address'

        result = scrape_website(excel_file, links, establishments, search_param)

        mock_driver.find_element.assert_called_with(By.ID, 'decrs_table')
        self.assertIsInstance(result, BytesIO)
