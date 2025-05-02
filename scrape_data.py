from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.core.os_manager import ChromeType
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from io import BytesIO
from rapidfuzz import fuzz
import pandas as pd
from read_in_establishments import convert_excel_to_csv, clean_text
import io

def get_driver():
    options = Options()
    options.add_argument("--headless")  # Run in headless mode
    options.add_argument("--no-sandbox")  # Needed for some CI environments
    options.add_argument("--disable-dev-shm-usage")  # Prevents issues with /dev/shm size
    options.add_argument("--disable-gpu")  # Safe for headless
    options.add_argument("--window-size=1920,1080")  # Ensure consistent viewport
    
    service = Service(ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install())

    driver = webdriver.Chrome(service=service, options=options)
    return driver

def scrape_website(excel_file, links, establishments_list, search_param):
    
    # Webdriver
    driver = get_driver()
    #Open excel file to write
    wb = load_workbook(excel_file)

    for link in links:
        # Open generated file
        driver.get(link)
        
        # Wait for elements to load
        driver.implicitly_wait(5)

        # Display all entries (ensures we look and compare all entries NOT just the first page if we have a partial match)
        entries_count_dropdown = driver.find_element("name", "decrs_table_length")
        select = Select(entries_count_dropdown)
        select.select_by_visible_text("All")
        
        # Entire table
        table = driver.find_element(By.ID, 'decrs_table')
        
        # Get all rows
        rows = table.find_elements(By.TAG_NAME, 'tr')
        
        # Extract data from each row
        for row in rows:
            try:
                name = row.find_element(By.CLASS_NAME, 'firm_name').text
                address = row.find_element(By.CLASS_NAME, 'decrs-address').text
                duns = row.find_element(By.CLASS_NAME, 'duns-number').text
                business = row.find_element(By.CLASS_NAME, 'business_operations').text
                expiration = row.find_element(By.CLASS_NAME, 'expiration_date').text
                #print(f'name: {name} DUNS: {duns}, Address: {address}')  

                if search_param == "address":
                    for i, original_address in enumerate(establishments_list):
                        clean_original_address = clean_text(original_address[1])
                        clean_scraped_address = clean_text(address)
                        match_ratio = fuzz.token_set_ratio(clean_original_address, clean_scraped_address)
                        if match_ratio >= 85:
                            print("Matched with a:", match_ratio, "for", original_address[1])
                            write_file(excel_file, wb, i, name, duns, business, expiration, search_param)
                            break
                elif search_param == "duns":
                    for i, original_duns in enumerate(establishments_list):
                        normalized_original_duns = original_duns[1].lstrip('0')
                        normalized_duns = duns.lstrip('0') #to handle cases where we have leading zeroes (sometimes ignored by excel sheets)
                        #print('normalized_og_duns:', normalized_original_duns, 'normalize_duns:', normalized_duns)
                        if normalized_original_duns == normalized_duns:
                            write_file(excel_file, wb, i, name, duns, business, expiration, search_param)
                            break
            except:
                continue  # Skip rows that don’t match

    # Save file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Close file and webdriver
    wb.close()
    driver.quit()

    return buffer


def write_file(excel_file, wb, count, name, duns, business, expiration, search_param):
    if isinstance(excel_file, BytesIO):
        df = pd.read_excel(excel_file)
    else:
        csv_file = convert_excel_to_csv(excel_file)
        df = pd.read_csv(csv_file)
    
    # Find column in which establishment specfied is
    establishment_label = f"Establishment {count+1}"
    column = df.columns.get_loc(df.columns[df.eq(establishment_label).any(axis=0)][0]) + 1

    # Find rows that need to be filled
    name_label = 'Establishment Name (DECRS)'
    name_row = df[df.eq(name_label).any(axis=1)].index[0] + 2
    
    duns_label = 'DUNS (DECRS)'
    duns_row = df[df.eq(duns_label).any(axis=1)].index[0] + 2
    
    business_label = 'Business Operations (DECRS)'
    business_row = df[df.eq(business_label).any(axis=1)].index[0] + 2
    
    expiration_label = 'Registration Expiration Date (DECRS)'
    expiration_row = df[df.eq(expiration_label).any(axis=1)].index[0] + 2

    if search_param=="address":
        wb.active = wb.sheetnames.index('Full Query') 
    elif search_param=="duns":
        wb.active = wb.sheetnames.index('Abbreviated Query')
    ws = wb.active

    ws.cell(row=name_row, column=column, value=name)
    ws.cell(row=duns_row, column=column, value=duns)
    ws.cell(row=business_row, column=column, value=business)
    ws.cell(row=expiration_row, column=column, value=expiration)

    print("Data successfully written to Excel!")